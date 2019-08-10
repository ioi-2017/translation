import os
from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from trans.models import Country, Language, User, Task, Contest
from trans.utils import get_trans_by_user_and_task
from openpyxl import load_workbook
from glob import glob


InitialDataFile = 'trans/initial_data/initial_data.xlsx'
TasksDirectory  = 'trans/initial_data/tasks/'


class Command(BaseCommand):
    help = 'Import initial data'

    def add_arguments(self, parser):
        parser.add_argument(
            '--reset',
            action='store_true',
            dest='reset',
            default=False,
            help='Remove previous data',
        )
        parser.add_argument(
            '--import', metavar='models', type=str, nargs='+',
            help='Select models to import',
            default=['countries', 'languages', 'users', 'tasks']
        )

    def handle(self, *args, **options):
        reset = options['reset']

        if options['import']:
            for entry in options['import']:
                if entry == 'languages':
                    self.import_languages(reset)
                if entry == 'countries':
                    self.import_countries(reset)
                if entry == 'users':
                    self.import_users(reset)
                if entry == 'tasks':
                    self.import_tasks(reset)

    def import_languages(self, reset):
        if reset:
            Language.objects.all().exclude(code__in=['en', 'fa']).delete()

        data = self.read_data('Languages', ['Language', 'Code', 'Direction'])

        for name, code, direction in data:
            Language.objects.get_or_create(name=name, code=code, rtl=(direction=='rtl'))
            # print(name, code, direction=='rtl')
        print('Languages imported.')

    def import_countries(self, reset):
        if reset:
            Country.objects.all().exclude(code__in=['IOI', 'ISC', 'TST']).delete()

        data = self.read_data('Countries', ['Country', 'Code'])

        for name, code in data:
            Country.objects.get_or_create(name=name, code=code)
            # print(name, code)
        print('Countries imported.')

    def import_users(self, reset):
        if reset:
            User.objects.filter(is_staff=False).all().exclude(country__code__in=['IOI', 'ISC']).delete()

        data = self.read_data('Users', ['Username', 'Country', 'Language', 'Password'])

        for username, country_code, language_code, password in data:
            country = Country.objects.get(code=country_code)
            language = Language.objects.get(code=language_code)
            user, created = User.objects.get_or_create(country=country, language=language, username=username)
            user.set_password(password)
            user.save()
            # print(country, language, username, password)
        print('Users imported.')

    def import_tasks(self, reset):
        if reset:
            Task.objects.all().delete()

        tasks_folder = os.path.join(settings.BASE_DIR, TasksDirectory)
        folders = glob(tasks_folder + '*/')

        for folder in folders:
            contest_slug = os.path.basename(os.path.normpath(folder))
            for file_name in glob(folder + '*.md'):
                task_name = os.path.basename(file_name).split('.')[0]
                task_parts = task_name.split('-')
                order = 1
                if len(task_parts) > 1:
                    order = int(task_parts[0])
                task_name = task_parts[-1]
                self.import_task(file_name, task_name, order, contest_slug)
                # print('Task {} imported to {}.'.format(task_name, contest_slug))
        print('Tasks imported.')

    def import_task(self, file_name, task_name, order, contest_slug):
        with open(file_name, 'r') as file:
            content = file.read()
            contest = Contest.objects.get(slug=contest_slug)
            task, created = Task.objects.get_or_create(name=task_name, order=order, contest=contest)
            user = User.objects.get(username="ISC")
            new_trans = get_trans_by_user_and_task(user, task)
            new_trans.add_version(content)
            if contest.public == True:
                task.publish_latest("Initial Release")

    def read_data(self, data_sheet, title_list):
        '''read data from spreadsheet'''
        data = []
        table = load_workbook(InitialDataFile)[data_sheet]
        titles = [c[0].value for c in table.columns]
        index = {t: titles.index(t) for t in title_list}
        for i in range(table.max_row - 1):
            row = table[i + 2]
            data.append(list([str(row[index[t]].value).strip() for t in title_list]))
        return data
